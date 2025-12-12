import aiohttp_jinja2
import jinja2
from aiohttp import web
import configparser
import asyncio
import aiohttp
import aiosqlite
import json
import os
import secrets
import time
import io
import math
import zipfile
from PIL import Image, ImageDraw, ImageFont
from urllib.parse import quote
from pypinyin import pinyin, Style
from dotenv import load_dotenv

load_dotenv()

DEFAULT_SETTINGS = {
    "title": "歌单",
    "live_url": "https://",
    "background_url": "/static/background.jpg",
    "singer_url": "/static/singer.jpg",
    "singer_name": "歌手名字",
    "singer_intro": "这里填写歌手介绍~",
}

class CaseSensitiveConfigParser(configparser.ConfigParser):
    def optionxform(self, optionstr):
        return optionstr


class Config:
    """负责读取和管理配置文件的类。"""

    def __init__(self, filename="config.ini"):
        self.config = CaseSensitiveConfigParser()
        self.config.read(filename)
        # cache env overrides
        self.env_admin_token = os.environ.get("QQZHU_ADMIN_TOKEN")

    def _get(self, section, key, default):
        if self.config.has_section(section) and key in self.config[section]:
            return self.config[section][key]
        return default

    def db_path(self):
        return os.environ.get("QQZHU_DB_PATH") or self._get("database", "path", "instance/qqzhu.db")

    def server_port(self):
        env_port = os.environ.get("QQZHU_PORT")
        if env_port:
            return int(env_port)
        return int(self._get("server", "port", 8080))

    def admin_token(self):
        return self.env_admin_token or self._get("server", "admin_token", "")


LOGIN_ATTEMPTS = {}
LOGIN_LIMIT = 5
LOGIN_WINDOW = 300  # seconds


def _login_attempts_info(ip: str):
    now = time.time()
    attempts = LOGIN_ATTEMPTS.get(ip, [])
    attempts = [t for t in attempts if now - t < LOGIN_WINDOW]
    LOGIN_ATTEMPTS[ip] = attempts
    remaining = max(LOGIN_LIMIT - len(attempts), 0)
    wait = 0
    if remaining == 0 and attempts:
        wait = max(0, int(LOGIN_WINDOW - (now - attempts[0])))
    return remaining, wait


def _add_login_failure(ip: str):
    now = time.time()
    attempts = LOGIN_ATTEMPTS.get(ip, [])
    attempts = [t for t in attempts if now - t < LOGIN_WINDOW]
    attempts.append(now)
    LOGIN_ATTEMPTS[ip] = attempts


def _reset_login_failure(ip: str):
    LOGIN_ATTEMPTS.pop(ip, None)


def update_env_var(key: str, value: str, env_path: str = ".env"):
    """Persist a key-value to .env, replacing if exists."""
    lines = []
    found = False
    if os.path.exists(env_path):
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                if line.strip().startswith(f"{key}="):
                    lines.append(f"{key}={value}\n")
                    found = True
                else:
                    lines.append(line)
    if not found:
        lines.append(f"{key}={value}\n")
    with open(env_path, "w", encoding="utf-8") as f:
        f.writelines(lines)
    # Also update process env for current runtime
    os.environ[key] = value


@web.middleware
async def csrf_middleware(request, handler):
    # Only enforce for admin-related endpoints
    is_admin_path = request.path.startswith("/admin")
    if is_admin_path:
        token = request.cookies.get("csrf_token") or secrets.token_urlsafe(32)
        request["csrf_token"] = token

        if request.method == "POST":
            try:
                form = await request.post()
            except Exception:
                form = {}
            form_token = form.get("csrf_token") if hasattr(form, "get") else None
            header_token = request.headers.get("X-CSRF-Token")
            cookie_token = request.cookies.get("csrf_token")
            final_token = form_token or header_token
            if not cookie_token or not final_token or final_token != cookie_token:
                raise web.HTTPForbidden(text="Invalid CSRF token")

    response = await handler(request)

    if is_admin_path and "csrf_token" in request:
        if request.cookies.get("csrf_token") != request["csrf_token"]:
            response.set_cookie("csrf_token", request["csrf_token"], httponly=True, samesite="Lax")
    return response


@web.middleware
async def security_headers_middleware(request, handler):
    response = await handler(request)
    response.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    # 轻量 CSP：允许自身与 https 资源，行内样式/脚本被现有页面使用，图片放宽到 https/data
    response.headers.setdefault(
        "Content-Security-Policy",
        "default-src 'self' https: data:; "
        "script-src 'self' 'unsafe-inline'; "
        "style-src 'self' 'unsafe-inline' https://cdnjs.cloudflare.com; "
        "img-src 'self' https: data:; "
        "font-src 'self' https://cdnjs.cloudflare.com data:; "
        "connect-src 'self' https:; "
        "frame-ancestors 'self';"
    )
    return response


async def create_db_connection(db_path):
    # Ensure the database directory exists before connecting
    db_dir = os.path.dirname(db_path)
    if db_dir:
        os.makedirs(db_dir, exist_ok=True)
    conn = await aiosqlite.connect(db_path)
    conn.row_factory = aiosqlite.Row
    await ensure_tables(conn)
    return conn


async def ensure_tables(conn):
    """Create required tables if they do not exist."""
    await conn.execute(
        """
        CREATE TABLE IF NOT EXISTS songs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            artist TEXT NOT NULL,
            language TEXT NOT NULL,
            genre TEXT NOT NULL,
            url TEXT NOT NULL
        )
        """
    )
    await ensure_settings_table(conn)
    await conn.commit()


async def ensure_settings_table(conn):
    await conn.execute(
        """
        CREATE TABLE IF NOT EXISTS site_settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
        """
    )
    await conn.commit()


def ensure_upload_dir():
    os.makedirs("static/uploads", exist_ok=True)


def save_file_field(file_field, prefix):
    """Save aiohttp FileField to static/uploads and return web path."""
    ensure_upload_dir()
    filename = file_field.filename
    if not filename:
        return None
    ext = os.path.splitext(filename)[1]
    safe_name = f"{prefix}_{int(time.time())}{ext}"
    path_fs = os.path.join("static", "uploads", safe_name)
    with open(path_fs, "wb") as f:
        content = file_field.file.read()
        f.write(content)
    return f"/static/uploads/{safe_name}"


def pick_font(font_path=None, size=38):
    """Pick a font that exists on the system."""
    if font_path and os.path.exists(font_path):
        try:
            return ImageFont.truetype(font_path, size)
        except Exception:
            pass
    try_fonts = [
        "msyhl.ttc",
        "SourceHanSansSC-Light.otf",
        "msyh.ttc",
        "simsun.ttc",
        "STHeiti Light.ttc",
    ]
    for tf in try_fonts:
        try:
            return ImageFont.truetype(tf, size)
        except Exception:
            continue
    font = ImageFont.load_default()
    font.size = size
    return font


def wrap_song_names(names, max_chars_per_line=35, max_songs_per_line=6):
    lines = []
    current = []
    current_chars = 0
    for name in names:
        name_len = len(name)
        if current and (current_chars + name_len > max_chars_per_line or len(current) >= max_songs_per_line):
            lines.append("  ".join(current))
            current = []
            current_chars = 0
        current.append(name)
        current_chars += name_len + 2
    if current:
        lines.append("  ".join(current))
    return lines


def combine_background_img(head_img, content_img, end_img, output_height):
    head_h = head_img.height
    content_h = content_img.height
    end_h = end_img.height
    width = head_img.width
    combined = Image.new("RGB", (width, output_height))
    combined.paste(head_img, (0, 0))
    content_area_h = output_height - head_h - end_h
    blocks = max(content_area_h // content_h, 0)
    for i in range(blocks + 1):
        y = head_h + i * content_h
        if y >= output_height - end_h:
            break
        combined.paste(content_img, (0, y))
    combined.paste(end_img, (0, output_height - end_h))
    return combined


def generate_playlist_image_from_bg(
    bg_bytes,
    content_start,
    end_start,
    names,
    font_path=None,
    font_size=38,
    max_chars_per_line=35,
    max_songs_per_line=6,
    line_height=None,
    output_dir="static/uploads",
):
    if not names:
        raise ValueError("歌曲列表为空")
    ensure_upload_dir()
    img = Image.open(io.BytesIO(bg_bytes))
    width, height = img.size
    if not (0 <= content_start < end_start <= height):
        raise ValueError("content_start/end_start 范围非法")

    head = img.crop((0, 0, width, content_start))
    content = img.crop((0, content_start, width, end_start))
    tail = img.crop((0, end_start, width, height))

    bg_color = content.getpixel((min(10, content.width - 1), min(10, content.height - 1)))
    text_color = (30, 30, 30) if sum(bg_color) > 382 else (240, 240, 240)

    lines = wrap_song_names(
        names,
        max_chars_per_line=max_chars_per_line,
        max_songs_per_line=max_songs_per_line,
    )
    resolved_line_height = max(line_height or content.height, 80)
    total_height = head.height + len(lines) * resolved_line_height + tail.height
    canvas = combine_background_img(head, content, tail, total_height)

    font = pick_font(font_path, font_size)
    draw = ImageDraw.Draw(canvas)
    shadow_color = (
        (bg_color[0] - 30, bg_color[1] - 30, bg_color[2] - 30)
        if sum(bg_color) > 382
        else (0, 0, 0)
    )

    y = head.height
    margin_left = 50
    for line in lines:
        text_w, text_h = draw.textsize(line, font=font)
        text_y = y + (resolved_line_height - text_h) // 2
        draw.text((margin_left + 2, text_y + 2), line, font=font, fill=shadow_color)
        draw.text((margin_left, text_y), line, font=font, fill=text_color)
        y += resolved_line_height

    ts = int(time.time())
    filename = f"playlist_{ts}.png"
    path_fs = os.path.join(output_dir, filename)
    canvas.save(path_fs, quality=95)
    return f"/static/uploads/{filename}"


def generate_playlist_pages_from_bg(
    bg_bytes,
    rect,
    names,
    font_path=None,
    font_size=38,
    max_chars_per_line=35,
    max_songs_per_line=6,
    line_height=80,
    output_dir="static/uploads",
):
    if not names:
        raise ValueError("歌曲列表为空")
    ensure_upload_dir()
    img = Image.open(io.BytesIO(bg_bytes)).convert("RGB")
    width, height = img.size
    x1, y1, x2, y2 = rect
    if not (0 <= x1 < x2 <= width and 0 <= y1 < y2 <= height):
        raise ValueError("矩形范围非法")

    lines = wrap_song_names(names, max_chars_per_line=max_chars_per_line, max_songs_per_line=max_songs_per_line)
    lines_per_page = max(1, (y2 - y1) // line_height)
    pages = math.ceil(len(lines) / lines_per_page)

    sample_x = min(max(x1 + 5, 0), width - 1)
    sample_y = min(max(y1 + 5, 0), height - 1)
    bg_color = img.getpixel((sample_x, sample_y))
    text_color = (30, 30, 30) if sum(bg_color) > 382 else (240, 240, 240)
    shadow_color = (
        (max(bg_color[0] - 30, 0), max(bg_color[1] - 30, 0), max(bg_color[2] - 30, 0))
        if sum(bg_color) > 382
        else (0, 0, 0)
    )
    font = pick_font(font_path, font_size)

    ts = int(time.time())
    saved_files = []
    for page in range(pages):
        start = page * lines_per_page
        end = start + lines_per_page
        page_lines = lines[start:end]
        canvas = img.copy()
        draw = ImageDraw.Draw(canvas)
        y = y1
        for line in page_lines:
            text_w, text_h = draw.textsize(line, font=font)
            text_y = y + (line_height - text_h) // 2
            draw.text((x1 + 2, text_y + 2), line, font=font, fill=shadow_color)
            draw.text((x1, text_y), line, font=font, fill=text_color)
            y += line_height
        filename = f"playlist_page_{ts}_{page+1:02d}.png"
        path_fs = os.path.join(output_dir, filename)
        canvas.save(path_fs, quality=95)
        saved_files.append(path_fs)

    zip_name = f"playlist_pages_{ts}.zip"
    zip_path = os.path.join(output_dir, zip_name)
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for fpath in saved_files:
            zf.write(fpath, arcname=os.path.basename(fpath))
    return {
        "zip_path": f"/static/uploads/{zip_name}",
        "files": [f"/static/uploads/{os.path.basename(p)}" for p in saved_files],
    }


def wants_json(request):
    accept = request.headers.get("Accept", "")
    return "application/json" in accept or request.headers.get("X-Requested-With") == "XMLHttpRequest"


async def fetch_songs(conn):
    """获取歌曲列表，按 id 升序。"""
    cursor = await conn.execute(
        "SELECT id, name, artist, language, genre, url FROM songs ORDER BY id ASC"
    )
    rows = await cursor.fetchall()
    await cursor.close()
    return [dict(row) for row in rows]


async def fetch_songs_sorted(conn):
    """获取歌曲列表，按原先规则排序（中文优先、短优先、拼音/字母）。"""
    songs = await fetch_songs(conn)

    def sort_key(song):
        name = song["name"]
        language = song["language"]
        language_priority = 0 if language == "中文" else 1
        word_count = len(name)
        if language == "中文":
            name_for_sort = "".join([item[0] for item in pinyin(name, style=Style.NORMAL)])
        else:
            name_for_sort = name.lower()
        return (language_priority, word_count, name_for_sort)

    return sorted(songs, key=sort_key)


async def fetch_unique_languages(conn):
    cursor = await conn.execute("SELECT DISTINCT language FROM songs")
    rows = await cursor.fetchall()
    await cursor.close()
    return [row["language"] for row in rows]


async def fetch_unique_genres(conn):
    cursor = await conn.execute("SELECT DISTINCT genre FROM songs")
    rows = await cursor.fetchall()
    await cursor.close()
    return [row["genre"] for row in rows]


async def get_settings(conn):
    await ensure_settings_table(conn)
    cursor = await conn.execute("SELECT key, value FROM site_settings")
    rows = await cursor.fetchall()
    await cursor.close()
    stored = {row["key"]: row["value"] for row in rows}
    merged = DEFAULT_SETTINGS.copy()
    merged.update({k: v for k, v in stored.items() if v is not None})
    return merged


async def update_settings(conn, settings: dict):
    await ensure_settings_table(conn)
    for key, value in settings.items():
        await conn.execute(
            "INSERT INTO site_settings (key, value) VALUES (?, ?) ON CONFLICT(key) DO UPDATE SET value = excluded.value",
            (key, value),
        )
    await conn.commit()


async def add_song(conn, name, artist, language, genre, url):
    cursor = await conn.execute(
        "INSERT INTO songs (name, artist, language, genre, url) VALUES (?, ?, ?, ?, ?)",
        (name, artist, language, genre, url),
    )
    song_id = cursor.lastrowid
    await conn.commit()
    await cursor.close()
    return song_id


async def update_song(conn, song_id, name, artist, language, genre, url):
    await conn.execute(
        """
        UPDATE songs
        SET name = ?, artist = ?, language = ?, genre = ?, url = ?
        WHERE id = ?
        """,
        (name, artist, language, genre, url, song_id),
    )
    await conn.commit()


async def delete_song(conn, song_id):
    await conn.execute("DELETE FROM songs WHERE id = ?", (song_id,))
    await conn.commit()


async def backup_songs(conn, dest_path):
    songs = await fetch_songs(conn)
    os.makedirs(os.path.dirname(dest_path), exist_ok=True)
    with open(dest_path, "w", encoding="utf-8") as f:
        json.dump(songs, f, ensure_ascii=False, indent=2)
    return dest_path


def parse_backup_json(text: str):
    data = json.loads(text)
    if not isinstance(data, list):
        raise ValueError("备份格式错误：根节点应为列表")
    for item in data:
        if not isinstance(item, dict):
            raise ValueError("备份格式错误：列表元素应为对象")
        for key in ("name", "artist", "language", "genre", "url"):
            if key not in item:
                raise ValueError(f"备份缺少必要字段: {key}")
    return data


async def restore_songs_from_data(conn, songs):
    # 清空并恢复
    await conn.execute("DELETE FROM songs")
    for s in songs:
        await add_song(
            conn,
            s.get("name", ""),
            s.get("artist", ""),
            s.get("language", ""),
            s.get("genre", ""),
            s.get("url", "-"),
        )
    await conn.commit()
    return len(songs)


def parse_backup_xlsx(data: bytes):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("未安装 openpyxl，无法解析 xlsx，请先安装 openpyxl")

    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    required = ["歌名", "歌手", "语言", "风格", "url"]
    if header[:5] != required:
        raise ValueError(f"xlsx 表头需为：{' | '.join(required)}")
    songs = []
    for row in rows[1:]:
        if row is None:
            continue
        vals = list(row) + [""] * (5 - len(row))
        name, artist, language, genre, url = ["" if v is None else str(v) for v in vals[:5]]
        if not name and not artist:
            continue
        songs.append(
            {
                "name": name,
                "artist": artist,
                "language": language,
                "genre": genre,
                "url": url or "-",
            }
        )
    return songs


def require_admin(request, form=None):
    """Check admin token via header/query/form/cookie; redirect to login if missing."""
    token_cfg = request.app["config"].admin_token()
    if not token_cfg:
        next_url = quote(str(request.rel_url))
        raise web.HTTPFound(location=f"/admin/setup?next={next_url}")
    provided = (
        request.headers.get("X-Admin-Token")
        or request.query.get("token")
        or (form.get("token") if form else None)
        or request.cookies.get("admin_token")
    )
    if provided == token_cfg:
        return provided
    next_url = quote(str(request.rel_url))
    raise web.HTTPFound(location=f"/admin/login?next={next_url}")


async def index(request):
    conn = request.app["db_conn"]
    songs = await fetch_songs_sorted(conn)
    languages = await fetch_unique_languages(conn)
    genres = await fetch_unique_genres(conn)
    settings = await get_settings(conn)
    return aiohttp_jinja2.render_template(
        "index.html",
        request,
        {"songs": songs, "languages": languages, "genres": genres, "settings": settings},
    )


async def proxy_image(request):
    image_url = request.query.get("url")
    if not image_url:
        raise web.HTTPBadRequest(text="Missing 'url' parameter")
    headers = {"Referer": "https://www.bilibili.com"}
    async with aiohttp.ClientSession() as session:
        async with session.get(image_url, headers=headers) as response:
            if response.status == 200:
                return web.Response(
                    body=await response.read(), content_type=response.headers["Content-Type"]
                )
            raise web.HTTPNotFound(text="Image not found")


async def admin_login_get(request):
    if not request.app["config"].admin_token():
        next_url = quote(str(request.rel_url))
        raise web.HTTPFound(location=f"/admin/setup?next={next_url}")
    return aiohttp_jinja2.render_template(
        "admin_login.html",
        request,
        {
            "next": request.query.get("next", "/admin"),
            "message": request.query.get("message", ""),
            "csrf_token": request.get("csrf_token", ""),
        },
    )


async def admin_login_post(request):
    ip = (request.headers.get("X-Forwarded-For", "").split(",")[0].strip() or request.remote or "unknown")
    remaining, wait = _login_attempts_info(ip)
    if remaining == 0:
        return aiohttp_jinja2.render_template(
            "admin_login.html",
            request,
            {
                "next": request.query.get("next", "/admin"),
                "message": f"尝试过多，请 {wait} 秒后再试",
                "csrf_token": request.get("csrf_token", ""),
            },
        )
    if not request.app["config"].admin_token():
        next_url = quote(str(request.rel_url))
        raise web.HTTPFound(location=f"/admin/setup?next={next_url}")
    form = await request.post()
    token_cfg = request.app["config"].admin_token()
    provided = form.get("token", "")
    next_url = form.get("next") or "/admin"
    if token_cfg and provided == token_cfg:
        resp = web.HTTPFound(location=next_url)
        resp.set_cookie("admin_token", provided, httponly=True, samesite="Lax")
        _reset_login_failure(ip)
        return resp
    _add_login_failure(ip)
    remaining_after, wait_after = _login_attempts_info(ip)
    message_text = "Token 错误或未设置"
    if remaining_after > 0:
        message_text += f"；还可再试 {remaining_after} 次"
    else:
        message_text += f"；请 {wait_after} 秒后再试"
    return aiohttp_jinja2.render_template(
        "admin_login.html",
        request,
        {"next": next_url, "message": message_text, "csrf_token": request.get("csrf_token", "")},
    )


async def admin_logout(request):
    resp = web.HTTPFound(location="/admin/login")
    resp.del_cookie("admin_token")
    return resp


async def admin_setup_get(request):
    """First-time token setup when no admin_token is configured."""
    if request.app["config"].admin_token():
        raise web.HTTPFound(location="/admin/login")
    return aiohttp_jinja2.render_template(
        "admin_setup.html",
        request,
        {
            "next": request.query.get("next", "/admin"),
            "message": request.query.get("message", ""),
            "csrf_token": request.get("csrf_token", ""),
        },
    )


async def admin_setup_post(request):
    if request.app["config"].admin_token():
        raise web.HTTPFound(location="/admin/login")
    form = await request.post()
    new_token = form.get("new_token", "").strip()
    confirm_token = form.get("confirm_token", "").strip()
    next_url = form.get("next") or "/admin"
    if not new_token:
        return aiohttp_jinja2.render_template(
            "admin_setup.html",
            request,
            {"next": next_url, "message": "Token 不能为空", "csrf_token": request.get("csrf_token", "")},
        )
    if new_token != confirm_token:
        return aiohttp_jinja2.render_template(
            "admin_setup.html",
            request,
            {"next": next_url, "message": "两次输入的 Token 不一致", "csrf_token": request.get("csrf_token", "")},
        )
    update_env_var("QQZHU_ADMIN_TOKEN", new_token)
    request.app["config"].env_admin_token = new_token
    resp = web.HTTPFound(location=next_url)
    resp.set_cookie("admin_token", new_token, httponly=True, samesite="Lax")
    return resp


async def admin_page(request):
    _ = require_admin(request)
    conn = request.app["db_conn"]
    songs = await fetch_songs(conn)
    settings = await get_settings(conn)
    return aiohttp_jinja2.render_template(
        "admin.html",
        request,
        {
            "songs": songs,
            "settings": settings,
            "message": request.query.get("message", ""),
            "token": request.query.get("token", ""),
            "env_admin_token": bool(request.app["config"].env_admin_token),
            "csrf_token": request.get("csrf_token", ""),
        },
    )


async def admin_action(request):
    conn = request.app["db_conn"]
    form = await request.post()
    require_admin(request, form)
    action = form.get("action")
    message = ""
    try:
        if action == "song_new":
            name = form.get("song_name", "").strip()
            artist = form.get("artist", "").strip()
            language = form.get("language", "").strip()
            genre = form.get("genre", "").strip()
            url = form.get("url", "").strip() or "-"
            new_id = await add_song(conn, name, artist, language, genre, url)
            message = "歌曲已添加"
            if wants_json(request):
                return web.json_response({"ok": True, "action": "song_new", "song": {
                    "id": new_id, "name": name, "artist": artist, "language": language, "genre": genre, "url": url
                }})
        elif action == "song_update":
            song_id = int(form.get("song_id", 0))
            name = form.get("song_name", "").strip()
            artist = form.get("artist", "").strip()
            language = form.get("language", "").strip()
            genre = form.get("genre", "").strip()
            url = form.get("url", "").strip() or "-"
            await update_song(conn, song_id, name, artist, language, genre, url)
            message = "歌曲已更新"
            if wants_json(request):
                return web.json_response({"ok": True, "action": "song_update", "song_id": song_id})
        elif action == "song_delete":
            song_id = int(form.get("song_id", 0))
            await delete_song(conn, song_id)
            message = "歌曲已删除"
            if wants_json(request):
                return web.json_response({"ok": True, "action": "song_delete", "song_id": song_id})
        elif action == "backup":
            dest = os.path.join("backup", "songs_backup.json")
            saved = await backup_songs(conn, dest)
            message = f"备份已生成: {saved}"
        elif action == "restore_backup":
            file_field = form.get("backup_file")
            try:
                if file_field and hasattr(file_field, "file") and file_field.filename:
                    filename = file_field.filename.lower()
                    if filename.endswith(".json"):
                        content = file_field.file.read().decode("utf-8")
                        songs = parse_backup_json(content)
                    elif filename.endswith(".xlsx"):
                        content = file_field.file.read()
                        songs = parse_backup_xlsx(content)
                    else:
                        raise ValueError("仅支持 .json 或 .xlsx 备份文件")
                    count = await restore_songs_from_data(conn, songs)
                    message = f"已从上传的备份恢复 {count} 首歌曲"
                    if wants_json(request):
                        return web.json_response({"ok": True, "action": "restore_backup", "count": count})
                else:
                    src = os.path.join("backup", "songs_backup.json")
                    if not os.path.exists(src):
                        raise FileNotFoundError(f"{src} 不存在")
                    with open(src, "r", encoding="utf-8") as f:
                        content = f.read()
                    songs = parse_backup_json(content)
                    count = await restore_songs_from_data(conn, songs)
                    message = f"已从本地备份恢复 {count} 首歌曲"
                    if wants_json(request):
                        return web.json_response({"ok": True, "action": "restore_backup", "count": count})
            except Exception as exc:
                message = f"恢复失败: {exc}"
        elif action == "settings":
            current_settings = await get_settings(conn)
            bg_file = form.get("background_file")
            singer_file = form.get("singer_file")
            new_settings = {
                "title": form.get("title", "").strip() or current_settings.get("title", DEFAULT_SETTINGS["title"]),
                "live_url": form.get("live_url", "").strip() or current_settings.get("live_url", DEFAULT_SETTINGS["live_url"]),
                "singer_name": form.get("singer_name", "").strip() or current_settings.get("singer_name", DEFAULT_SETTINGS["singer_name"]),
                "singer_intro": form.get("singer_intro", "").strip() or current_settings.get("singer_intro", DEFAULT_SETTINGS["singer_intro"]),
                "background_url": current_settings.get("background_url", DEFAULT_SETTINGS["background_url"]),
                "singer_url": current_settings.get("singer_url", DEFAULT_SETTINGS["singer_url"]),
            }
            if hasattr(bg_file, "file") and bg_file.filename:
                saved = save_file_field(bg_file, "bg")
                if saved:
                    new_settings["background_url"] = saved
            if hasattr(singer_file, "file") and singer_file.filename:
                saved = save_file_field(singer_file, "singer")
                if saved:
                    new_settings["singer_url"] = saved
            await update_settings(conn, new_settings)
            message = "站点信息已更新"
        elif action == "update_admin_token":
            new_token = form.get("new_token", "").strip()
            confirm_token = form.get("confirm_token", "").strip()
            if not new_token:
                raise ValueError("新 token 不能为空")
            if new_token != confirm_token:
                raise ValueError("两次输入的 token 不一致")
            update_env_var("QQZHU_ADMIN_TOKEN", new_token)
            request.app["config"].env_admin_token = new_token
            message = "admin_token 已更新，请重新登录"
            if wants_json(request):
                return web.json_response({"ok": True, "action": "update_admin_token", "message": message})
        elif action == "generate_playlist_image":
            bg_field = form.get("bg_image")
            if not (bg_field and hasattr(bg_field, "file") and bg_field.filename):
                raise ValueError("请上传背景图")
            try:
                content_start = int(form.get("content_start", 0))
                end_start = int(form.get("end_start", 0))
            except Exception:
                raise ValueError("content_start/end_start 需要是整数")
            songs_sorted = await fetch_songs_sorted(conn)
            names = [s["name"] for s in songs_sorted]
            font_path = form.get("font_path", "").strip() or None
            font_size = int(form.get("font_size", 38) or 38)
            line_height = form.get("line_height")
            line_height_val = int(line_height) if line_height else None
            max_chars_per_line = int(form.get("max_chars_per_line", 35) or 35)
            max_songs_per_line = int(form.get("max_songs_per_line", 6) or 6)
            generated_path = generate_playlist_image_from_bg(
                bg_field.file.read(),
                content_start,
                end_start,
                names,
                font_path=font_path,
                font_size=font_size,
                max_chars_per_line=max_chars_per_line,
                max_songs_per_line=max_songs_per_line,
                line_height=line_height_val,
            )
            if wants_json(request):
                return web.json_response({"ok": True, "action": "generate_playlist_image", "path": generated_path})
            # For normal form submit, stream the generated file as download
            fs_path = generated_path.lstrip("/\\")
            fs_path = fs_path if os.path.isabs(fs_path) else os.path.join(os.getcwd(), fs_path)
            if not os.path.exists(fs_path):
                raise web.HTTPNotFound(text="生成的长图不存在")
            filename = os.path.basename(fs_path)
            headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
            return web.FileResponse(path=fs_path, headers=headers)
        elif action == "generate_playlist_pages":
            bg_field = form.get("bg_image_small")
            if not (bg_field and hasattr(bg_field, "file") and bg_field.filename):
                raise ValueError("请上传背景图")
            try:
                x1 = int(form.get("rect_x1", 0))
                y1 = int(form.get("rect_y1", 0))
                x2 = int(form.get("rect_x2", 0))
                y2 = int(form.get("rect_y2", 0))
            except Exception:
                raise ValueError("矩形坐标需要是整数")
            font_size = int(form.get("font_size", 38) or 38)
            line_height = int(form.get("line_height", 80) or 80)
            max_chars_per_line = int(form.get("max_chars_per_line", 35) or 35)
            max_songs_per_line = int(form.get("max_songs_per_line", 6) or 6)
            songs_sorted = await fetch_songs_sorted(conn)
            names = [s["name"] for s in songs_sorted]
            font_path = form.get("font_path", "").strip() or None
            result = generate_playlist_pages_from_bg(
                bg_field.file.read(),
                (x1, y1, x2, y2),
                names,
                font_path=font_path,
                font_size=font_size,
                max_chars_per_line=max_chars_per_line,
                max_songs_per_line=max_songs_per_line,
                line_height=line_height,
            )
            zip_path = result["zip_path"]
            if wants_json(request):
                return web.json_response({"ok": True, "action": "generate_playlist_pages", "zip_path": zip_path, "files": result["files"]})
            fs_path = zip_path.lstrip("/\\")
            fs_path = fs_path if os.path.isabs(fs_path) else os.path.join(os.getcwd(), fs_path)
            if not os.path.exists(fs_path):
                raise web.HTTPNotFound(text="生成的文件不存在")
            filename = os.path.basename(fs_path)
            headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
            return web.FileResponse(path=fs_path, headers=headers)
        else:
            message = "未知操作"
    except Exception as exc:
        message = f"操作失败: {exc}"
        if wants_json(request):
            return web.json_response({"ok": False, "message": message})

    if wants_json(request):
        return web.json_response({"ok": True, "message": message})
    params = [f"message={quote(message)}"]
    return web.HTTPFound(location="/admin?" + "&".join(params) + "#songs")


async def admin_download_backup(request):
    """Generate and send latest backup file for download."""
    _ = require_admin(request)
    conn = request.app["db_conn"]
    dest = os.path.join("backup", "songs_backup.json")
    saved = await backup_songs(conn, dest)
    headers = {"Content-Disposition": 'attachment; filename="songs_backup.json"'}
    return web.FileResponse(path=saved, headers=headers)


async def init_app():
    app = web.Application(
        client_max_size=10 * 1024 * 1024,
        middlewares=[csrf_middleware, security_headers_middleware],
    )
    aiohttp_jinja2.setup(app, loader=jinja2.FileSystemLoader("templates"))

    app.router.add_static("/static/", path="static", name="static")
    app.router.add_get("/proxy-image", proxy_image)
    app.router.add_get("/healthz", lambda request: web.Response(text="ok"))

    config = Config("config.ini")
    db_conn = await create_db_connection(config.db_path())
    app["db_conn"] = db_conn
    app["config"] = config

    app.router.add_get("/", index)
    app.router.add_get("/admin/login", admin_login_get)
    app.router.add_post("/admin/login", admin_login_post)
    app.router.add_get("/admin/logout", admin_logout)
    app.router.add_get("/admin/setup", admin_setup_get)
    app.router.add_post("/admin/setup", admin_setup_post)
    app.router.add_get("/admin", admin_page)
    app.router.add_post("/admin/action", admin_action)
    app.router.add_get("/admin/download-backup", admin_download_backup)

    async def close_db(app):
        await app["db_conn"].close()

    app.on_cleanup.append(close_db)
    return app


if __name__ == "__main__":
    port = Config("config.ini").server_port()
    web.run_app(init_app(), port=port)
